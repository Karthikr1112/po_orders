import csv
import json
import logging
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.views.decorators.http import require_POST

from .models import (
    Buyer,
    PurchaseOrder,
    PurchaseOrderItem,
    Season,
    SubCategory,
    SubCategorySize,
    Vendor,
    SubCategoryPriceRange,
)

logger = logging.getLogger("po_sheet")


# ---------------------------------------------------------------------------
# Batch query helpers  (eliminate N+1 patterns throughout the module)
# ---------------------------------------------------------------------------

def _batch_pr_aggregates(subcat_ids):
    """
    Return a dict: {subcategory_id: {'total_amt': Decimal, 'total_qty': int}}
    for all price ranges belonging to the given subcategory IDs.
    Single query regardless of how many IDs are passed.
    """
    if not subcat_ids:
        return {}
    rows = (
        SubCategoryPriceRange.objects
        .filter(subcategory_id__in=subcat_ids)
        .values("subcategory_id")
        .annotate(total_amt=Sum("approved_amount"), total_qty=Sum("approved_quantity"))
    )
    return {r["subcategory_id"]: r for r in rows}


def _batch_placed_maps(subcat_ids, extra_filter=None):
    """
    Return (placed_amt_map, placed_qty_map) dicts keyed by subcategory_id,
    scoped to submitted (non-draft) POs.  Optional extra_filter is a Q object
    applied to the PurchaseOrderItem queryset (e.g. buyer / season).
    Single pair of queries regardless of how many IDs are passed.
    """
    if not subcat_ids:
        return {}, {}
    qs = PurchaseOrderItem.objects.filter(
        purchase_order__is_draft=False, subcategory_id__in=subcat_ids
    )
    if extra_filter:
        qs = qs.filter(extra_filter)
    placed_amt = {
        r["subcategory_id"]: r["total"]
        for r in qs.values("subcategory_id").annotate(total=Sum("tot_amt"))
    }
    placed_qty = {
        r["subcategory_id"]: r["total"]
        for r in qs.values("subcategory_id").annotate(total=Sum("tot_qty"))
    }
    return placed_amt, placed_qty


# ---------------------------------------------------------------------------
# Local Price Range helper
# ---------------------------------------------------------------------------

def get_or_sync_subcategory_price_ranges(subcategory):
    """
    Fetch price ranges for the subcategory from the local database.
    If none exist, create a default 0.00 to 0.00 range so that the user
    can enter/edit buying price ranges.
    """
    ranges = list(SubCategoryPriceRange.objects.filter(subcategory=subcategory).order_by('sales_from_range'))
    if not ranges:
        obj, created = SubCategoryPriceRange.objects.get_or_create(
            subcategory=subcategory,
            sales_from_range=Decimal("0.00"),
            sales_to_range=Decimal("0.00"),
            defaults={
                "buying_from_range": Decimal("0.00"),
                "buying_to_range": Decimal("0.00"),
            }
        )
        ranges = [obj]
    return ranges


# ---------------------------------------------------------------------------
# MSSQL helpers
# ---------------------------------------------------------------------------

def _get_mssql_conn():
    """Open a fresh pyodbc connection from settings. Always close in a finally block."""
    from django.conf import settings
    import pyodbc

    db = settings.DATABASES["mssql"]
    driver = db.get("OPTIONS", {}).get("driver", "ODBC Driver 17 for SQL Server")
    server = db.get("HOST", "localhost")
    port = db.get("PORT", "")
    server_str = f"{server},{port}" if port else server
    conn_str = (
        f"DRIVER={{{driver}}};SERVER={server_str};"
        f"DATABASE={db.get('NAME','')};UID={db.get('USER','')};PWD={db.get('PASSWORD','')}"
    )
    return pyodbc.connect(conn_str, timeout=10)


def fetch_mssql_vendors(query=None):
    vendors_list = []
    conn = None
    try:
        conn = _get_mssql_conn()
        cursor = conn.cursor()
        if query:
            like_val = f"%{query}%"
            cursor.execute(
                "SELECT TOP 100 lifnr, name, city, gstNo, postCode, street, email1"
                " FROM [dbo].[Vendor_B]"
                " WHERE lifnr LIKE ? OR name LIKE ? OR city LIKE ?",
                (like_val, like_val, like_val),
            )
        else:
            cursor.execute(
                "SELECT TOP 50 lifnr, name, city, gstNo, postCode, street, email1"
                " FROM [dbo].[Vendor_B]"
            )
        for row in cursor.fetchall():
            vendors_list.append({
                "vendor_code": row[0],
                "vendor_name": row[1] or "",
                "address": row[5] or "",
                "city": row[2] or "",
                "state": "",
                "pin_code": str(row[4]) if row[4] is not None else "",
                "email": row[6] or "",
                "phone": "",
                "gst_number": row[3] or "",
                "contact_person": "",
            })
    except Exception:
        logger.exception("MSSQL error in fetch_mssql_vendors")
    finally:
        if conn:
            conn.close()
    return vendors_list


def get_or_create_mssql_vendor(vendor_code):
    if not vendor_code:
        return None
    vendor_code = str(vendor_code).strip()

    local = Vendor.objects.filter(vendor_code=vendor_code).first()
    if local:
        return local

    conn = None
    try:
        conn = _get_mssql_conn()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name, city, gstNo, postCode, street, email1"
            " FROM [dbo].[Vendor_B] WHERE lifnr = ?",
            (vendor_code,),
        )
        row = cursor.fetchone()
        if row:
            vendor, _ = Vendor.objects.get_or_create(
                vendor_code=vendor_code,
                defaults={
                    "vendor_name": row[0] or vendor_code,
                    "city": row[1] or "",
                    "gst_number": row[2] or "",
                    "pin_code": str(row[3]) if row[3] is not None else "",
                    "address": row[4] or "",
                    "email": row[5] or "",
                    "state": "",
                    "phone": "",
                    "contact_person": "",
                },
            )
            return vendor
    except Exception:
        logger.exception("MSSQL error in get_or_create_mssql_vendor: %s", vendor_code)
    finally:
        if conn:
            conn.close()

    # Fallback: create a stub vendor so the PO can still be saved
    vendor, _ = Vendor.objects.get_or_create(
        vendor_code=vendor_code,
        defaults={"vendor_name": vendor_code},
    )
    return vendor


# ---------------------------------------------------------------------------
# Budget helpers
# ---------------------------------------------------------------------------

def get_subcategory_remaining_budget(subcategory, exclude_po_id=None, include_draft_po_id=None):
    if not subcategory:
        return Decimal("0")

    # Total approved = sum of all price range approved_amounts for this subcategory
    agg = SubCategoryPriceRange.objects.filter(subcategory=subcategory).aggregate(
        total=Sum("approved_amount")
    )
    approved_total = agg["total"] or Decimal("0")
    if not approved_total:
        return Decimal("0")

    qs = PurchaseOrderItem.objects.filter(
        subcategory=subcategory,
        purchase_order__is_draft=False,
    )
    if exclude_po_id:
        qs = qs.exclude(purchase_order_id=exclude_po_id)

    spent = qs.aggregate(total=Sum("tot_amt"))["total"] or Decimal("0")

    if include_draft_po_id:
        draft_spent = (
            PurchaseOrderItem.objects.filter(
                purchase_order_id=include_draft_po_id,
                purchase_order__is_draft=True,
                subcategory=subcategory,
            ).aggregate(total=Sum("tot_amt"))["total"]
            or Decimal("0")
        )
        spent += draft_spent

    return approved_total - spent


# ---------------------------------------------------------------------------
# PO number generation  (race-safe: unique constraint is the final guard)
# ---------------------------------------------------------------------------

def generate_next_po_number():
    last = (
        PurchaseOrder.objects.filter(po_number__regex=r"^PO-\d+$")
        .order_by("-id")
        .values_list("po_number", flat=True)
        .first()
    )
    if last:
        m = re.search(r"(\d+)$", last)
        next_num = int(m.group(1)) + 1 if m else 1
    else:
        next_num = 1
    return f"PO-{next_num:04d}"


# ---------------------------------------------------------------------------
# PO totals helper  (used inline where on_commit is too late for the response)
# ---------------------------------------------------------------------------

def _refresh_po_totals(po):
    agg = po.items.aggregate(total_qty=Sum("tot_qty"), grand_total=Sum("tot_amt"))
    PurchaseOrder.objects.filter(pk=po.pk).update(
        total_quantity=agg["total_qty"] or 0,
        grand_total=agg["grand_total"] or Decimal("0"),
    )
    po.total_quantity = agg["total_qty"] or 0
    po.grand_total = agg["grand_total"] or Decimal("0")


# ---------------------------------------------------------------------------
# Auth views
# ---------------------------------------------------------------------------

def login_view(request):
    if request.user.is_authenticated:
        return redirect("po_sheet")
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST.get("username", "").strip(),
            password=request.POST.get("password", ""),
        )
        if user is not None:
            login(request, user)
            return redirect("po_sheet")
        messages.error(request, "Invalid username or password")
    return render(request, "po_sheet/login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


# ---------------------------------------------------------------------------
# Main PO sheet
# ---------------------------------------------------------------------------

@login_required
@ensure_csrf_cookie
def po_sheet(request):
    user = request.user
    po = PurchaseOrder.objects.filter(created_by=user, is_draft=True).first()
    if not po:
        po = PurchaseOrder.objects.create(
            created_by=user,
            is_draft=True,
            po_number=generate_next_po_number(),
        )

    # Ensure draft has a proper PO-XXXX number
    if not re.match(r"^PO-\d{4,}$", po.po_number):
        po.po_number = generate_next_po_number()
        po.save(update_fields=["po_number"])

    po_items = po.items.select_related("subcategory").all()

    agg = po_items.aggregate(total_qty=Sum("tot_qty"), grand_total=Sum("tot_amt"))
    totals = {
        "items": agg["total_qty"] or 0,
        "subtotal": Decimal("0"),
        "grand_total": agg["grand_total"] or Decimal("0"),
    }

    # Keep stored totals in sync (cheap: only 2 columns)
    PurchaseOrder.objects.filter(pk=po.pk).update(
        total_quantity=totals["items"],
        grand_total=totals["grand_total"],
    )

    # Budget availability per subcategory — 2 queries total, no N+1
    subcat_ids = [item.subcategory_id for item in po_items]
    pr_agg_map = _batch_pr_aggregates(subcat_ids)
    placed_map, placed_qty_map = _batch_placed_maps(subcat_ids)

    for item in po_items:
        sc = item.subcategory
        pr = pr_agg_map.get(sc.id, {})
        approved = pr.get("total_amt") or Decimal("0")
        approved_qty = pr.get("total_qty") or 0
        sc.has_budget = approved > 0
        placed = placed_map.get(sc.id, Decimal("0"))
        sc.available_budget = float(max(Decimal("0"), approved - Decimal(str(placed))))
        sc.approved_amount_val = float(approved)
        sc.approved_quantity_val = approved_qty
        sc.spent_amount_val = float(placed)
        sc.spent_qty_val = int(placed_qty_map.get(sc.id, 0))

    context = {
        "po": po,
        "po_items": po_items,
        "totals": totals,
        "buyers": Buyer.objects.all(),
        "seasons": Season.objects.all(),
        "schedules_json": json.dumps(po.delivery_schedules or []),
        "categories": [],
        "all_subcategories": [],
        "vendors": [],
    }
    return render(request, "po_sheet/po_sheet.html", context)


# ---------------------------------------------------------------------------
# Vendor search / select
# ---------------------------------------------------------------------------

@login_required
def search_vendor(request):
    query = (request.GET.get("q") or "").strip()
    vendors = fetch_mssql_vendors(query)
    results = [
        {
            "id": v["vendor_code"],
            "code": v["vendor_code"],
            "name": v["vendor_name"],
            "contact": v["contact_person"],
            "phone": v["phone"],
            "email": v["email"],
            "address": v["address"],
            "city": v["city"],
            "state": v["state"],
            "pin_code": v["pin_code"],
            "gst_number": v["gst_number"],
        }
        for v in vendors
    ]
    return JsonResponse({"vendors": results})


@login_required
@require_POST
def select_vendor(request):
    try:
        data = json.loads(request.body)
        vendor_id = (data.get("vendor_id") or "").strip()
        if not vendor_id:
            return JsonResponse({"success": False, "error": "Vendor is required"})
        vendor = get_or_create_mssql_vendor(vendor_id)
        if not vendor:
            return JsonResponse({"success": False, "error": "Vendor not found"})
        po = PurchaseOrder.objects.filter(created_by=request.user, is_draft=True).first()
        if not po:
            return JsonResponse({"success": False, "error": "Draft PO not found"})
        po.vendor = vendor
        po.save(update_fields=["vendor"])
        return JsonResponse({"success": True})
    except Exception:
        logger.exception("select_vendor error")
        return JsonResponse({"success": False, "error": "Server error"})


@login_required
def vendor_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = Vendor.objects.all()
    if q:
        qs = qs.filter(
            Q(vendor_code__icontains=q)
            | Q(vendor_name__icontains=q)
            | Q(city__icontains=q)
        )
    return render(request, "po_sheet/vendors.html", {"vendors": qs[:100], "q": q})


# ---------------------------------------------------------------------------
# Subcategory search (MSSQL)
# ---------------------------------------------------------------------------

@login_required
def search_subcategory(request):
    query = (request.GET.get("q") or "").strip()
    buyer_id = request.GET.get("buyer", "").strip()

    try:
        if buyer_id:
            try:
                buyer_id_int = int(buyer_id)
            except ValueError:
                buyer_id_int = None

            if buyer_id_int:
                qs = SubCategory.objects.filter(buyers__id=buyer_id_int)
                if query:
                    qs = qs.filter(name__icontains=query)

                subcat_list = list(qs[:50])
                subcat_ids = [sc.id for sc in subcat_list]

                # 2 queries for all price-range + placed data (no N+1)
                pr_agg_map = _batch_pr_aggregates(subcat_ids)
                placed_map, placed_qty_map = _batch_placed_maps(subcat_ids)

                results = []
                for sc in subcat_list:
                    pr = pr_agg_map.get(sc.id, {})
                    approved_amount = pr.get("total_amt") or Decimal("0")
                    approved_quantity = pr.get("total_qty") or 0
                    has_budget = approved_amount > 0
                    placed = placed_map.get(sc.id, Decimal("0"))
                    available_budget = float(
                        max(Decimal("0"), approved_amount - Decimal(str(placed)))
                    )
                    results.append({
                        "id": sc.id,
                        "name": sc.name,
                        "category": "Local Category",
                        "ch4_code": sc.ch4_code or "",
                        "available_budget": available_budget,
                        "has_budget": has_budget,
                        "approved_amount": float(approved_amount),
                        "approved_quantity": approved_quantity,
                        "spent_amount": float(placed),
                        "spent_quantity": int(placed_qty_map.get(sc.id, 0)),
                        "unit_price": 0.0,
                    })
                return JsonResponse({"subcategories": results})

        # --- MSSQL branch ---
        rows = []
        conn = None
        try:
            conn = _get_mssql_conn()
            cursor = conn.cursor()
            if query:
                term = f"%{query}%"
                cursor.execute(
                    "SELECT DISTINCT TOP 100 SubCategory, CH4"
                    " FROM [dbo].[MCH_View]"
                    " WHERE SubCategory IS NOT NULL AND (SubCategory LIKE ? OR CH4 LIKE ?)",
                    (term, term),
                )
            else:
                cursor.execute(
                    "SELECT DISTINCT TOP 50 SubCategory, CH4"
                    " FROM [dbo].[MCH_View] WHERE SubCategory IS NOT NULL"
                )
            rows = cursor.fetchall()
        except Exception:
            logger.exception("MSSQL error in search_subcategory")
            rows = []
        finally:
            if conn:
                conn.close()

        if not rows:
            return JsonResponse({"subcategories": []})

        # Single query for all local subcategories in the MSSQL result set
        names = [r[0] for r in rows]
        local_map = {
            sc.name: sc
            for sc in SubCategory.objects.filter(name__in=names)
        }

        # Batch price-range + placed data for local subcategories only (no N+1)
        local_ids = [sc.id for sc in local_map.values()]
        pr_agg_map = _batch_pr_aggregates(local_ids)
        placed_map, placed_qty_map = _batch_placed_maps(local_ids)

        results = []
        for subcat_name, ch4 in rows:
            sc = local_map.get(subcat_name)
            has_budget = False
            approved_amount = Decimal("0")
            approved_quantity = 0
            available_budget = 0.0
            spent_placed = Decimal("0")
            subcat_id = subcat_name  # fallback when not yet in MySQL

            if sc:
                subcat_id = sc.id
                pr = pr_agg_map.get(sc.id, {})
                approved_amount = pr.get("total_amt") or Decimal("0")
                approved_quantity = pr.get("total_qty") or 0
                has_budget = approved_amount > 0
                spent_placed = placed_map.get(sc.id, Decimal("0"))
                available_budget = float(
                    max(Decimal("0"), approved_amount - Decimal(str(spent_placed)))
                )

            results.append({
                "id": subcat_id,
                "name": subcat_name,
                "category": "MSSQL Category",
                "ch4_code": ch4 or "",
                "available_budget": available_budget,
                "has_budget": has_budget,
                "approved_amount": float(approved_amount),
                "approved_quantity": approved_quantity,
                "spent_amount": float(spent_placed),
                "spent_quantity": int(placed_qty_map.get(sc.id, 0)) if sc else 0,
                "unit_price": 0.0,
            })

        return JsonResponse({"subcategories": results})

    except Exception:
        logger.exception("search_subcategory error")
        return JsonResponse({"subcategories": [], "error": "Server error"}, status=500)


# ---------------------------------------------------------------------------
# PO item helpers
# ---------------------------------------------------------------------------

def _item_to_dict(item):
    """Serialize a PurchaseOrderItem for JSON responses."""
    return {
        "id": item.id,
        "name": item.subcategory.name,
        # SubCategory has no FK to Category — use ch4_code as the identifier
        "category": item.subcategory.ch4_code or "",
        "ch4_code": item.subcategory.ch4_code or "",
        "order_qty": item.order_qty,
        "tot_qty": item.tot_qty,
        "unit_price": float(item.unit_price),
        "discount_percentage": float(item.discount_percentage),
        "tot_amt": float(item.tot_amt),
        "item_type": item.item_type,
    }


def _po_totals_dict(po, agg):
    return {
        "total_items": po.total_quantity,
        "subtotal": float(agg.get("subtotal") or 0),
        "grand_total": float(po.grand_total),
        "remaining_budget": 0,  # vendor-level budget removed; subcategory-level used instead
    }


# ---------------------------------------------------------------------------
# Add / update / delete PO items
# ---------------------------------------------------------------------------

@login_required
@require_POST
def add_po_item(request):
    try:
        data = json.loads(request.body)
        subcategory_id = data.get("subcategory_id")
        quantity = max(0, int(data.get("quantity", 1)))
        unit_price = Decimal(str(data.get("unit_price") or "0"))
        discount_percentage = Decimal(str(data.get("discount_percentage") or "0"))
        item_type = data.get("item_type", "Fresh")

        po = get_object_or_404(PurchaseOrder, created_by=request.user, is_draft=True)

        try:
            subcategory = SubCategory.objects.get(id=subcategory_id)
        except (SubCategory.DoesNotExist, ValueError, TypeError):
            subcategory, _ = SubCategory.objects.get_or_create(name=str(subcategory_id))

        item, created = PurchaseOrderItem.objects.get_or_create(
            purchase_order=po,
            subcategory=subcategory,
            defaults={
                "unit_price": unit_price,
                "order_qty": quantity,
                "discount_percentage": discount_percentage,
                "item_type": item_type,
            },
        )
        if not created:
            item.order_qty = quantity
            item.discount_percentage = discount_percentage
            item.unit_price = unit_price
            item.item_type = item_type
            item.save()

        _refresh_po_totals(po)

        po_items = po.items.select_related("subcategory").all()
        agg = po_items.aggregate(
            subtotal=Sum(F("tot_qty") * F("unit_price")),
        )
        return JsonResponse({
            "success": True,
            "items": [_item_to_dict(it) for it in po_items],
            "totals": _po_totals_dict(po, agg),
        })
    except (InvalidOperation, ValueError) as e:
        return JsonResponse({"success": False, "error": f"Invalid number: {e}"})
    except Exception:
        logger.exception("add_po_item error")
        return JsonResponse({"success": False, "error": "Server error"})


@login_required
@require_POST
def update_po_item(request, item_id):
    try:
        data = json.loads(request.body)
        item = get_object_or_404(
            PurchaseOrderItem,
            id=item_id,
            purchase_order__created_by=request.user,
            purchase_order__is_draft=True,
        )

        if "quantity" in data:
            item.order_qty = max(0, int(data["quantity"]))
        if "discount_percentage" in data:
            item.discount_percentage = Decimal(str(data["discount_percentage"]))
        if "unit_price" in data:
            item.unit_price = Decimal(str(data["unit_price"]))
        if "item_type" in data:
            item.item_type = data["item_type"]
        item.save()

        po = item.purchase_order
        _refresh_po_totals(po)

        agg = po.items.aggregate(subtotal=Sum(F("tot_qty") * F("unit_price")))
        return JsonResponse({
            "success": True,
            "item": _item_to_dict(item),
            "totals": _po_totals_dict(po, agg),
        })
    except (InvalidOperation, ValueError) as e:
        return JsonResponse({"success": False, "error": f"Invalid number: {e}"})
    except Exception:
        logger.exception("update_po_item error")
        return JsonResponse({"success": False, "error": "Server error"})


@login_required
def delete_po_item(request, item_id):
    if request.method != "DELETE":
        return JsonResponse({"success": False, "error": "Invalid method"})
    try:
        item = get_object_or_404(
            PurchaseOrderItem,
            id=item_id,
            purchase_order__created_by=request.user,
            purchase_order__is_draft=True,
        )
        po = item.purchase_order
        item.delete()
        _refresh_po_totals(po)

        agg = po.items.aggregate(subtotal=Sum(F("tot_qty") * F("unit_price")))
        return JsonResponse({
            "success": True,
            "totals": _po_totals_dict(po, agg),
        })
    except Exception:
        logger.exception("delete_po_item error")
        return JsonResponse({"success": False, "error": "Server error"})


@login_required
@require_POST
def add_manual_item(request):
    try:
        data = json.loads(request.body)
        item_name = (data.get("item_name") or "").strip()
        unit_price = float(data.get("unit_price", 0))
        quantity = max(0, int(data.get("quantity", 1)))

        if not item_name:
            return JsonResponse({"success": False, "error": "Item name is required"})
        if unit_price <= 0:
            return JsonResponse({"success": False, "error": "Price must be greater than 0"})

        subcategory, _ = SubCategory.objects.get_or_create(name=item_name)

        po = PurchaseOrder.objects.filter(created_by=request.user, is_draft=True).first()
        if not po:
            po = PurchaseOrder.objects.create(
                created_by=request.user,
                is_draft=True,
                po_number=generate_next_po_number(),
            )

        existing = po.items.filter(subcategory=subcategory).first()
        if existing:
            existing.order_qty += quantity
            existing.unit_price = Decimal(str(unit_price))
            existing.save()
        else:
            PurchaseOrderItem.objects.create(
                purchase_order=po,
                subcategory=subcategory,
                order_qty=quantity,
                unit_price=Decimal(str(unit_price)),
            )

        _refresh_po_totals(po)

        po_items = po.items.select_related("subcategory").all()
        agg = po_items.aggregate(subtotal=Sum(F("tot_qty") * F("unit_price")))
        return JsonResponse({
            "success": True,
            "items": [_item_to_dict(it) for it in po_items],
            "totals": _po_totals_dict(po, agg),
        })
    except Exception:
        logger.exception("add_manual_item error")
        return JsonResponse({"success": False, "error": "Server error"})


# ---------------------------------------------------------------------------
# PO field updates
# ---------------------------------------------------------------------------

@login_required
@require_POST
def update_po_fields(request):
    try:
        data = json.loads(request.body)
        po = PurchaseOrder.objects.filter(created_by=request.user, is_draft=True).first()
        if not po:
            return JsonResponse({"success": False, "error": "No draft PO found"})

        update_fields = []
        if "po_type" in data:
            po.po_type = data["po_type"]; update_fields.append("po_type")
        if "season" in data:
            po.season = data["season"]; update_fields.append("season")
        if "agent" in data:
            po.agent = data["agent"]; update_fields.append("agent")
        if "notes" in data:
            po.notes = data["notes"]; update_fields.append("notes")
        # Do NOT allow clients to override po_number freely — it must be system-generated
        if "buyer" in data:
            buyer_id = data["buyer"]
            po.buyer = Buyer.objects.filter(id=buyer_id).first() if buyer_id else None
            update_fields.append("buyer")
        if "vendor" in data:
            vendor_id = data["vendor"]
            po.vendor = get_or_create_mssql_vendor(vendor_id) if vendor_id else None
            update_fields.append("vendor")
        if "po_date" in data and data["po_date"]:
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                try:
                    po.po_date = datetime.strptime(data["po_date"], fmt).date()
                    update_fields.append("po_date")
                    break
                except ValueError:
                    pass

        if update_fields:
            po.save(update_fields=update_fields)
        return JsonResponse({"success": True})
    except Exception:
        logger.exception("update_po_fields error")
        return JsonResponse({"success": False, "error": "Server error"})


# ---------------------------------------------------------------------------
# Save PO (atomic bulk save + submit)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def save_po(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"})

    try:
        with transaction.atomic():
            po = (
                PurchaseOrder.objects
                .select_for_update()
                .filter(created_by=request.user, is_draft=True)
                .first()
            )
            if not po:
                return JsonResponse({"success": False, "error": "No draft PO found"})

            # Ensure a proper sequential number
            if not re.match(r"^PO-\d+$", po.po_number):
                po.po_number = generate_next_po_number()

            # PO header fields
            buyer_id = data.get("buyer")
            po.buyer = Buyer.objects.filter(id=buyer_id).first() if buyer_id else None
            po.agent = data.get("agent", "")
            po.notes = data.get("notes", "")
            po.delivery_schedules = data.get("delivery_schedules", [])
            po.po_type = data.get("po_type", "")
            po.season = data.get("season", "")

            po_date_str = data.get("po_date", "")
            if po_date_str:
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                    try:
                        po.po_date = datetime.strptime(po_date_str, fmt).date()
                        break
                    except ValueError:
                        pass

            items_data = data.get("items", [])

            # --- Validate Admin Budget / Qty Limits per Price Range ---
            base_items = PurchaseOrderItem.objects.filter(purchase_order__is_draft=False)
            if po.buyer:
                base_items = base_items.filter(purchase_order__buyer=po.buyer)
            if po.season:
                base_items = base_items.filter(purchase_order__season=po.season)



            proposed_totals = {}
            for item in items_data:
                subcat_id = item.get("subcategory_id")
                if not subcat_id:
                    continue
                try:
                    subcat = SubCategory.objects.get(id=subcat_id)
                except (SubCategory.DoesNotExist, ValueError, TypeError):
                    subcat, _ = SubCategory.objects.get_or_create(name=str(subcat_id))

                unit_price = Decimal(str(item.get("unit_price", 0)))
                qty = int(item.get("order_qty", 0))
                discount_percentage = Decimal(str(item.get("discount_percentage", 0)))
                discount_factor = Decimal("1") - (discount_percentage / Decimal("100"))
                tot_amt = unit_price * Decimal(qty) * discount_factor

                price_ranges = list(subcat.price_ranges.all())
                if not price_ranges:
                    price_ranges = get_or_sync_subcategory_price_ranges(subcat)

                matched_range = None
                for pr in price_ranges:
                    if pr.buying_from_range <= unit_price <= pr.buying_to_range:
                        matched_range = pr
                        break

                if not matched_range:
                    raise ValueError(
                        f"No price range defined for '{subcat.name}' that covers the unit price ₹{unit_price:,.2f}."
                    )

                approved_amt = matched_range.approved_amount or Decimal("0.00")
                approved_qty = matched_range.approved_quantity or 0

                if approved_amt == Decimal("0.00") and approved_qty == 0:
                    raise ValueError(
                        f"Saving not allowed. '{subcat.name}' in range [₹{matched_range.buying_from_range:,.2f} - ₹{matched_range.buying_to_range:,.2f}] "
                        f"does not have an approved budget or quantity set by the admin."
                    )

                key = (subcat.id, matched_range.id)
                if key not in proposed_totals:
                    proposed_totals[key] = {
                        "qty": 0,
                        "amount": Decimal("0.00"),
                        "subcategory_name": subcat.name,
                        "range_str": f"₹{matched_range.buying_from_range} - ₹{matched_range.buying_to_range}",
                        "approved_amount": approved_amt,
                        "approved_quantity": approved_qty,
                        "range_obj": matched_range,
                    }
                proposed_totals[key]["qty"] += qty
                proposed_totals[key]["amount"] += tot_amt

            for (subcat_id, range_id), info in proposed_totals.items():
                pr = info["range_obj"]
                subcat_name = info["subcategory_name"]
                range_str = info["range_str"]

                submitted_qs = base_items.filter(
                    subcategory_id=subcat_id,
                    unit_price__gte=pr.buying_from_range,
                    unit_price__lte=pr.buying_to_range,
                )

                agg_spent = submitted_qs.aggregate(
                    spent_amt=Sum("tot_amt"),
                    spent_qty=Sum("tot_qty")
                )
                spent_amt = agg_spent["spent_amt"] or Decimal("0.00")
                spent_qty = agg_spent["spent_qty"] or 0

                approved_amt = info["approved_amount"]
                approved_qty = info["approved_quantity"]
                proposed_amt = info["amount"]
                proposed_qty = info["qty"]



                if approved_amt > Decimal("0.00"):
                    total_amt = spent_amt + proposed_amt
                    if total_amt > approved_amt:
                        raise ValueError(f"Budget limit reached for '{subcat_name}' in range [{range_str}]. Approved: ₹{approved_amt:,.2f}, Spent so far: ₹{spent_amt:,.2f}, This PO: ₹{proposed_amt:,.2f}. Exceeds budget by ₹{total_amt - approved_amt:,.2f}.")

                if approved_qty > 0:
                    total_qty = spent_qty + proposed_qty
                    if total_qty > approved_qty:
                        raise ValueError(f"Quantity limit reached for '{subcat_name}' in range [{range_str}]. Approved: {approved_qty}, Placed so far: {spent_qty}, This PO: {proposed_qty}. Exceeds quantity limit by {total_qty - approved_qty}.")

            # --- Replace all items in one shot ---
            # Disconnect signals during bulk-insert so on_commit isn't triggered
            # per item — we'll do one aggregate update at the end instead.
            from django.db.models.signals import post_delete, post_save
            from po_sheet.signals import (
                update_po_totals_on_item_delete,
                update_po_totals_on_item_save,
            )
            post_save.disconnect(update_po_totals_on_item_save, sender=PurchaseOrderItem)
            post_delete.disconnect(update_po_totals_on_item_delete, sender=PurchaseOrderItem)
            try:
                po.items.all().delete()
                new_items = []
                for item in items_data:
                    subcat_id = item.get("subcategory_id")
                    if not subcat_id:
                        continue
                    try:
                        subcat = SubCategory.objects.get(id=subcat_id)
                    except (SubCategory.DoesNotExist, ValueError, TypeError):
                        subcat, _ = SubCategory.objects.get_or_create(name=str(subcat_id))

                    po_item = PurchaseOrderItem(
                        purchase_order=po,
                        subcategory=subcat,
                        unit_price=Decimal(str(item.get("unit_price", 0))),
                        order_qty=int(item.get("order_qty", 0)),
                        discount_percentage=Decimal(str(item.get("discount_percentage", 0))),
                        item_type=item.get("item_type", "Fresh"),
                        size_allocations=item.get("size_allocations", {}),
                    )
                    # Call model's save() so tot_qty/tot_amt are computed
                    po_item.save()
                    new_items.append(po_item)
            finally:
                # Always reconnect — even if an exception occurred
                post_save.connect(update_po_totals_on_item_save, sender=PurchaseOrderItem)
                post_delete.connect(update_po_totals_on_item_delete, sender=PurchaseOrderItem)

            # Mark submitted and persist header
            po.is_draft = False
            po.save()

            # Final aggregate (single query)
            agg = po.items.aggregate(
                total_qty=Sum("tot_qty"),
                grand_total=Sum("tot_amt"),
                subtotal=Sum(F("tot_qty") * F("unit_price")),
            )
            PurchaseOrder.objects.filter(pk=po.pk).update(
                total_quantity=agg["total_qty"] or 0,
                grand_total=agg["grand_total"] or Decimal("0"),
            )

            return JsonResponse({
                "success": True,
                "po_number": po.po_number,
                "totals": {
                    "total_items": agg["total_qty"] or 0,
                    "subtotal": float(agg["subtotal"] or 0),
                    "grand_total": float(agg["grand_total"] or 0),
                },
            })
    except ValueError as e:
        return JsonResponse({"success": False, "error": str(e)})
    except Exception:
        logger.exception("save_po error")
        return JsonResponse({"success": False, "error": "Server error — PO not saved"})


# ---------------------------------------------------------------------------
# Submit PO (lightweight path — just validates & flips is_draft)
# ---------------------------------------------------------------------------

@login_required
def submit_po(request):
    po = get_object_or_404(PurchaseOrder, created_by=request.user, is_draft=True)

    if not po.vendor:
        messages.error(request, "Please select a vendor before submitting")
        return redirect("po_sheet")
    if not po.items.exists():
        messages.error(request, "Please add at least one item before submitting")
        return redirect("po_sheet")

    PurchaseOrder.objects.filter(pk=po.pk).update(is_draft=False)
    messages.success(request, f"Purchase Order {po.po_number} submitted successfully")
    return redirect("all_records")


# ---------------------------------------------------------------------------
# New / delete PO
# ---------------------------------------------------------------------------

@login_required
def new_po(request):
    PurchaseOrder.objects.filter(created_by=request.user, is_draft=True).delete()
    PurchaseOrder.objects.create(
        po_number=generate_next_po_number(),
        created_by=request.user,
        is_draft=True,
    )
    messages.success(request, "New purchase order created")
    return redirect("po_sheet")


@login_required
def delete_po(request, po_id):
    po = get_object_or_404(PurchaseOrder, id=po_id)
    if not request.user.is_staff and po.created_by != request.user:
        messages.error(request, "You do not have permission to delete this Purchase Order.")
        return redirect("all_records")
    po_number = po.po_number
    po.delete()
    messages.success(request, f"Purchase Order {po_number} deleted.")
    return redirect("all_records")


# ---------------------------------------------------------------------------
# Preview / print PO
# ---------------------------------------------------------------------------

@login_required
def preview_po(request):
    po_id = request.GET.get("po_id")
    if po_id:
        try:
            po = PurchaseOrder.objects.get(id=po_id)
            # Permission: non-staff can only preview their own POs
            if not request.user.is_staff and po.created_by != request.user:
                return render(request, "po_sheet/print_po.html", {"error": "Permission denied"})
        except (PurchaseOrder.DoesNotExist, ValueError):
            po = None
    else:
        po = (
            PurchaseOrder.objects.filter(created_by=request.user, is_draft=True).first()
            or PurchaseOrder.objects.filter(created_by=request.user).order_by("-updated_at").first()
        )

    if not po:
        return render(request, "po_sheet/print_po.html", {"error": "No PO found"})

    po_items = po.items.select_related("subcategory").all()
    num_items = len(po_items)
    extra_rows_count = max(0, 12 - num_items)
    extra_rows = range(num_items + 1, num_items + extra_rows_count + 1)
    
    agg = po.items.aggregate(subtotal=Sum(F("tot_qty") * F("unit_price")))
    subtotal = agg["subtotal"] or Decimal("0")
    now = datetime.now()
    return render(request, "po_sheet/print_po.html", {
        "po": po,
        "po_items": po_items,
        "extra_rows": extra_rows,
        "subtotal": subtotal,
        "today": now.strftime("%d/%m/%Y"),
        "print_time": now.strftime("%d/%m/%Y, %I:%M:%S %p"),
    })


# ---------------------------------------------------------------------------
# Send PO (stub — wire up real email/WhatsApp here)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def send_po(request):
    try:
        data = json.loads(request.body)
        email = data.get("email", "")
        whatsapp = data.get("whatsapp", "")
        # TODO: integrate Django email backend or WhatsApp API
        logger.info("send_po stub: email=%s whatsapp=%s", email, whatsapp)
        return JsonResponse({
            "success": True,
            "message": f"PO sent to {email or 'N/A'} / {whatsapp or 'N/A'} (stub)",
        })
    except Exception:
        logger.exception("send_po error")
        return JsonResponse({"success": False, "error": "Server error"})


# ---------------------------------------------------------------------------
# All records
# ---------------------------------------------------------------------------

@login_required
def all_records(request):
    qs = (
        PurchaseOrder.objects
        .select_related("vendor", "created_by", "buyer")
        .prefetch_related("items__subcategory")
        .filter(is_draft=False)
        .order_by("-created_at")
    )
    if not request.user.is_staff:
        qs = qs.filter(created_by=request.user)

    search = (request.GET.get("q") or "").strip()
    if search:
        qs = qs.filter(
            Q(po_number__icontains=search)
            | Q(vendor__vendor_name__icontains=search)
            | Q(vendor__vendor_code__icontains=search)
            | Q(created_by__username__icontains=search)
        )

    vendor_filter = (request.GET.get("vendor") or "").strip()
    if vendor_filter:
        qs = qs.filter(vendor__vendor_code=vendor_filter)

    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()
    if date_from:
        qs = qs.filter(po_date__gte=date_from)
    if date_to:
        qs = qs.filter(po_date__lte=date_to)

    # Paginate to avoid loading all POs at once
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get("page"))


    vendors_qs = (
        PurchaseOrder.objects.filter(is_draft=False)
        .exclude(vendor__isnull=True)
        .values("vendor__vendor_code", "vendor__vendor_name")
        .distinct()
        .order_by("vendor__vendor_name")
    )

    return render(request, "po_sheet/all_records.html", {
        "records": page_obj.object_list,
        "page_obj": page_obj,
        "search": search,
        "vendor_filter": vendor_filter,
        "date_from": date_from,
        "date_to": date_to,
        "vendors": vendors_qs,
        "total": paginator.count,
        "is_staff": request.user.is_staff,
    })


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

@login_required
def export_po_csv(request, po_id):
    po = get_object_or_404(PurchaseOrder, id=po_id)
    if not request.user.is_staff and po.created_by != request.user:
        messages.error(request, "You do not have permission to export this Purchase Order.")
        return redirect("all_records")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{po.po_number}_Export.csv"'
    writer = csv.writer(response)

    writer.writerow([
        "PO Number", "PO Date", "PO Type", "Season", "Buyer", "Agent",
        "Vendor Name", "Vendor Code", "Total Quantity", "Grand Total",
        "Status", "Notes", "Delivery Schedules", "Created By", "Created At", "Updated At",
    ])
    writer.writerow([
        po.po_number, po.po_date, po.po_type, po.season,
        po.buyer.name if po.buyer else "",
        po.agent,
        po.vendor.vendor_name if po.vendor else "",
        po.vendor.vendor_code if po.vendor else "",
        po.total_quantity, po.grand_total,
        "Draft" if po.is_draft else "Saved",
        po.notes, po.delivery_schedules,
        po.created_by.username if po.created_by else "",
        po.created_at.strftime("%Y-%m-%d %H:%M:%S") if po.created_at else "",
        po.updated_at.strftime("%Y-%m-%d %H:%M:%S") if po.updated_at else "",
    ])
    writer.writerow([])
    writer.writerow([
        "Item #", "SubCategory", "CH4 Code", "Item Type",
        "Order Qty", "Total Qty", "Unit Price", "Discount %", "Total Amount", "Size Allocations",
    ])
    for i, item in enumerate(po.items.select_related("subcategory").all(), 1):
        sizes = item.size_allocations or {}
        writer.writerow([
            i,
            item.subcategory.name if item.subcategory else "",
            item.subcategory.ch4_code if item.subcategory else "",
            item.item_type,
            item.order_qty, item.tot_qty,
            item.unit_price, item.discount_percentage, item.tot_amt,
            " | ".join(f"{k}:{v}" for k, v in sizes.items()),
        ])
    return response


# ---------------------------------------------------------------------------
# Admin budget
# ---------------------------------------------------------------------------

@login_required
def admin_budget(request):
    if not request.user.is_staff:
        messages.error(request, "Only admin can manage budgets")
        return redirect("po_sheet")

    # Buyer filter
    buyer_id = request.GET.get("buyer", "").strip() or request.POST.get("buyer", "").strip()
    selected_buyer = None
    all_buyers = Buyer.objects.order_by("name")
    if buyer_id:
        try:
            selected_buyer = Buyer.objects.get(id=buyer_id)
        except (Buyer.DoesNotExist, ValueError):
            selected_buyer = None

    # Season filter
    season_id = request.GET.get("season", "").strip() or request.POST.get("season", "").strip()
    selected_season = None
    all_seasons = Season.objects.order_by("name")
    if season_id:
        try:
            selected_season = Season.objects.get(id=season_id)
        except (Season.DoesNotExist, ValueError):
            selected_season = None

    if request.method == "POST":
        # All budget edits are handled via AJAX (/update-price-range/).
        # POST just redirects back, preserving filters.
        url = "/admin-budget/"
        params = []
        if buyer_id:
            params.append(f"buyer={buyer_id}")
        if season_id:
            params.append(f"season={season_id}")
        if params:
            url += "?" + "&".join(params)
        return redirect(url)

    # Base PO items (submitted only)
    base_items = PurchaseOrderItem.objects.filter(purchase_order__is_draft=False)
    if selected_buyer:
        base_items = base_items.filter(purchase_order__buyer=selected_buyer)
    if selected_season:
        base_items = base_items.filter(purchase_order__season=selected_season.name)

    # We want to group PO items by subcategory
    po_items_by_subcat = {}
    for item in base_items.select_related("purchase_order", "subcategory"):
        po_items_by_subcat.setdefault(item.subcategory_id, []).append(item)

    # Spent per subcategory (respects buyer and season filters)
    placed_map = {
        row["subcategory_id"]: row["total"]
        for row in base_items.values("subcategory_id").annotate(total=Sum("tot_amt"))
    }

    # Budget rows — show all subcategories for selected buyer, only when buyer is selected
    if selected_buyer:
        subcat_qs = list(
            SubCategory.objects
            .filter(buyers=selected_buyer)
            .prefetch_related("buyers", "price_ranges")
            .order_by("name")
        )
    else:
        subcat_qs = []

    # Ensure every subcategory has at least one price range (batch: only one extra
    # query per subcategory that is genuinely missing ranges, usually zero).
    subcat_ids_all = [sc.id for sc in subcat_qs]
    existing_range_ids = set(
        SubCategoryPriceRange.objects
        .filter(subcategory_id__in=subcat_ids_all)
        .values_list("subcategory_id", flat=True)
        .distinct()
    )
    for sc in subcat_qs:
        if sc.id not in existing_range_ids:
            get_or_sync_subcategory_price_ranges(sc)

    # Refresh prefetch so newly created ranges are visible
    if subcat_ids_all:
        from django.db.models import Prefetch
        pr_prefetch = SubCategoryPriceRange.objects.filter(
            subcategory_id__in=subcat_ids_all
        ).order_by("sales_from_range")
        pr_by_sc = {}
        for pr in pr_prefetch:
            pr_by_sc.setdefault(pr.subcategory_id, []).append(pr)

    rows = []
    for sc in subcat_qs:
        spent = placed_map.get(sc.id, Decimal("0"))
        buyer_names = [b.name for b in sc.buyers.all()]

        price_ranges = pr_by_sc.get(sc.id, [])

        # Match each item to a price range based on buying price
        items = po_items_by_subcat.get(sc.id, [])
        range_spent_amount = {}  # range_id -> Decimal
        range_spent_qty = {}     # range_id -> int
        
        for item in items:
            matched_range = None
            for pr in price_ranges:
                if pr.buying_from_range <= item.unit_price <= pr.buying_to_range:
                    matched_range = pr
                    break
            if matched_range:
                range_spent_amount[matched_range.id] = range_spent_amount.get(matched_range.id, Decimal("0.00")) + item.tot_amt
                range_spent_qty[matched_range.id] = range_spent_qty.get(matched_range.id, 0) + item.tot_qty

        for pr in price_ranges:
            pr.spent_amount = range_spent_amount.get(pr.id, Decimal("0.00"))
            pr.spent_quantity = range_spent_qty.get(pr.id, 0)
            pr.balance = (pr.approved_amount or Decimal("0.00")) - pr.spent_amount

        # Total approved = sum of price range budgets (single source of truth)
        approved = sum((pr.approved_amount or Decimal("0.00")) for pr in price_ranges)
        approved_qty = sum((pr.approved_quantity or 0) for pr in price_ranges)

        # Serialize to JSON for data attributes in the template
        ranges_list = []
        for r in price_ranges:
            ranges_list.append({
                "id": r.id,
                "sales_from": float(r.sales_from_range),
                "sales_to": float(r.sales_to_range),
                "buying_from": float(r.buying_from_range),
                "buying_to": float(r.buying_to_range),
                "approved_amount": float(r.approved_amount),
                "approved_quantity": r.approved_quantity,
                "spent_amount": float(r.spent_amount),
                "spent_quantity": r.spent_quantity,
                "balance": float(r.balance),
            })
        import json
        price_ranges_json = json.dumps(ranges_list)

        rows.append({
            "id": sc.id,
            "name": sc.name,
            "ch4_code": sc.ch4_code or "—",
            "approved_amount": approved,
            "approved_quantity": approved_qty,
            "spent_amount": spent,
            "balance_amount": approved - spent,
            "buyers": buyer_names,
            "price_ranges": price_ranges,
            "price_ranges_json": price_ranges_json,
        })

    paginator = Paginator(rows, 30)
    page_obj = paginator.get_page(request.GET.get("page"))
    return render(request, "po_sheet/admin_budget.html", {
        "subcategories": page_obj.object_list,
        "page_obj": page_obj,
        "total_budgeted_count": paginator.count,
        "all_buyers": all_buyers,
        "selected_buyer": selected_buyer,
        "buyer_id": buyer_id,
        "all_seasons": all_seasons,
        "selected_season": selected_season,
        "season_id": season_id,
    })


@login_required
def budget_spent_details(request):
    if not request.user.is_staff:
        messages.error(request, "Only admin can view budget details")
        return redirect("po_sheet")

    # Filters
    buyer_id = request.GET.get("buyer", "").strip()
    selected_buyer = None
    all_buyers = Buyer.objects.order_by("name")
    if buyer_id:
        try:
            selected_buyer = Buyer.objects.get(id=buyer_id)
        except (Buyer.DoesNotExist, ValueError):
            selected_buyer = None

    season_id = request.GET.get("season", "").strip()
    selected_season = None
    all_seasons = Season.objects.order_by("name")
    if season_id:
        try:
            selected_season = Season.objects.get(id=season_id)
        except (Season.DoesNotExist, ValueError):
            selected_season = None

    if not selected_buyer:
        return render(request, "po_sheet/budget_spent_details.html", {
            "rows": [],
            "all_buyers": all_buyers,
            "selected_buyer": None,
            "buyer_id": "",
            "all_seasons": all_seasons,
            "season_id": season_id,
            "grand_total_approved_amount": Decimal("0.00"),
            "grand_total_approved_quantity": 0,
            "grand_total_spent_amount": Decimal("0.00"),
            "grand_total_spent_quantity": 0,
            "grand_total_balance": Decimal("0.00"),
        })

    # Fetch PO items for non-draft POs
    po_items_qs = PurchaseOrderItem.objects.filter(purchase_order__is_draft=False).select_related("purchase_order", "subcategory")
    po_items_qs = po_items_qs.filter(purchase_order__buyer=selected_buyer)
    if selected_season:
        po_items_qs = po_items_qs.filter(purchase_order__season=selected_season.name)

    # We want to group PO items by subcategory
    po_items_by_subcat = {}
    for item in po_items_qs:
        po_items_by_subcat.setdefault(item.subcategory_id, []).append(item)

    subcat_qs = SubCategory.objects.filter(buyers=selected_buyer).prefetch_related("price_ranges").order_by("name")

    rows = []
    grand_total_approved_amount = Decimal("0.00")
    grand_total_approved_qty = 0
    grand_total_spent_amount = Decimal("0.00")
    grand_total_spent_qty = 0
    grand_total_balance = Decimal("0.00")

    for sc in subcat_qs:
        price_ranges = list(sc.price_ranges.all())  # uses prefetch cache
        if not price_ranges:
            price_ranges = get_or_sync_subcategory_price_ranges(sc)
            
        items = po_items_by_subcat.get(sc.id, [])

        # Match each item to a price range based on buying price
        range_spent_amount = {}  # range_id -> Decimal
        range_spent_qty = {}     # range_id -> int
        
        unclassified_amount = Decimal("0.00")
        unclassified_qty = 0

        for item in items:
            # We match to a price range where buying_from_range <= item.unit_price <= buying_to_range
            matched_range = None
            for pr in price_ranges:
                if pr.buying_from_range <= item.unit_price <= pr.buying_to_range:
                    matched_range = pr
                    break
            
            if matched_range:
                range_spent_amount[matched_range.id] = range_spent_amount.get(matched_range.id, Decimal("0.00")) + item.tot_amt
                range_spent_qty[matched_range.id] = range_spent_qty.get(matched_range.id, 0) + item.tot_qty
            else:
                unclassified_amount += item.tot_amt
                unclassified_qty += item.tot_qty

        # Construct ranges list
        sc_ranges_data = []
        sc_total_approved = Decimal("0.00")
        sc_total_approved_qty = 0
        sc_total_spent = Decimal("0.00")
        sc_total_spent_qty = 0

        for pr in price_ranges:
            approved_amt = pr.approved_amount or Decimal("0.00")
            approved_q = pr.approved_quantity or 0
            
            spent_amt = range_spent_amount.get(pr.id, Decimal("0.00"))
            spent_q = range_spent_qty.get(pr.id, 0)
            
            bal = approved_amt - spent_amt
            
            sc_total_approved += approved_amt
            sc_total_approved_qty += approved_q
            sc_total_spent += spent_amt
            sc_total_spent_qty += spent_q

            sc_ranges_data.append({
                "sales_from": pr.sales_from_range,
                "sales_to": pr.sales_to_range,
                "buying_from": pr.buying_from_range,
                "buying_to": pr.buying_to_range,
                "approved_amount": approved_amt,
                "approved_quantity": approved_q,
                "spent_amount": spent_amt,
                "spent_quantity": spent_q,
                "balance": bal,
            })

        if unclassified_amount > 0 or unclassified_qty > 0:
            sc_total_spent += unclassified_amount
            sc_total_spent_qty += unclassified_qty
            sc_ranges_data.append({
                "is_unclassified": True,
                "approved_amount": Decimal("0.00"),
                "approved_quantity": 0,
                "spent_amount": unclassified_amount,
                "spent_quantity": unclassified_qty,
                "balance": -unclassified_amount,
            })

        sc_balance = sc_total_approved - sc_total_spent
        
        grand_total_approved_amount += sc_total_approved
        grand_total_approved_qty += sc_total_approved_qty
        grand_total_spent_amount += sc_total_spent
        grand_total_spent_qty += sc_total_spent_qty
        grand_total_balance += sc_balance

        rows.append({
            "id": sc.id,
            "name": sc.name,
            "ranges": sc_ranges_data,
            "total_approved_amount": sc_total_approved,
            "total_approved_quantity": sc_total_approved_qty,
            "total_spent_amount": sc_total_spent,
            "total_spent_quantity": sc_total_spent_qty,
            "balance": sc_balance,
        })

    return render(request, "po_sheet/budget_spent_details.html", {
        "rows": rows,
        "all_buyers": all_buyers,
        "selected_buyer": selected_buyer,
        "buyer_id": buyer_id,
        "all_seasons": all_seasons,
        "selected_season": selected_season,
        "season_id": season_id,
        "grand_total_approved_amount": grand_total_approved_amount,
        "grand_total_approved_quantity": grand_total_approved_qty,
        "grand_total_spent_amount": grand_total_spent_amount,
        "grand_total_spent_quantity": grand_total_spent_qty,
        "grand_total_balance": grand_total_balance,
    })


@login_required
@require_POST
def update_price_range(request):
    try:
        data = json.loads(request.body)
        range_id = data.get("range_id")

        updates = {}
        if "sales_from" in data:
            updates["sales_from_range"] = Decimal(str(data["sales_from"]))
        if "sales_to" in data:
            updates["sales_to_range"] = Decimal(str(data["sales_to"]))
        if "buying_from" in data:
            updates["buying_from_range"] = Decimal(str(data["buying_from"]))
        if "buying_to" in data:
            updates["buying_to_range"] = Decimal(str(data["buying_to"]))
        if "approved_amount" in data:
            updates["approved_amount"] = Decimal(str(data["approved_amount"]))
        if "approved_quantity" in data:
            updates["approved_quantity"] = int(data["approved_quantity"])

        if updates:
            SubCategoryPriceRange.objects.filter(id=range_id).update(**updates)

        return JsonResponse({"success": True})
    except Exception as e:
        logger.error(f"Error in update_price_range API: {e}")
        return JsonResponse({"success": False, "error": str(e)})


@login_required
@require_POST
def add_price_range(request):
    try:
        data = json.loads(request.body)
        subcategory_id = data.get("subcategory_id")
        subcat = SubCategory.objects.get(id=subcategory_id)
        
        # Find maximum sales_to_range for this subcategory to avoid duplicate key constraint
        max_range = SubCategoryPriceRange.objects.filter(subcategory=subcat).order_by("-sales_to_range").first()
        if max_range:
            new_sales_from = max_range.sales_to_range + Decimal("1.00")
            new_sales_to = new_sales_from
        else:
            new_sales_from = Decimal("0.00")
            new_sales_to = Decimal("0.00")
            
        new_range = SubCategoryPriceRange.objects.create(
            subcategory=subcat,
            sales_from_range=new_sales_from,
            sales_to_range=new_sales_to,
            buying_from_range=Decimal("0.00"),
            buying_to_range=Decimal("0.00")
        )

        return JsonResponse({
            "success": True,
            "range_id": new_range.id,
            "sales_from": float(new_sales_from),
            "sales_to": float(new_sales_to),
            "buying_from": 0,
            "buying_to": 0,
            "approved_amount": 0,
            "approved_quantity": 0,
        })
    except Exception as e:
        logger.error(f"Error in add_price_range API: {e}")
        return JsonResponse({"success": False, "error": str(e)})


@login_required
@require_POST
def delete_price_range(request):
    try:
        data = json.loads(request.body)
        range_id = data.get("range_id")
        SubCategoryPriceRange.objects.filter(id=range_id).delete()
        return JsonResponse({"success": True})
    except Exception as e:
        logger.error(f"Error in delete_price_range API: {e}")
        return JsonResponse({"success": False, "error": str(e)})




@login_required
def get_subcategory_ranges(request, subcategory_id):
    try:
        sc = SubCategory.objects.get(id=subcategory_id)
    except (SubCategory.DoesNotExist, ValueError):
        sc = SubCategory.objects.filter(name=subcategory_id).first()
    if not sc:
        return JsonResponse({"success": True, "ranges": []})

    price_ranges = get_or_sync_subcategory_price_ranges(sc)

    # Fetch ALL submitted items for this subcategory in ONE query, then
    # classify in-memory — avoids N per-range DB queries.
    all_items = list(
        PurchaseOrderItem.objects.filter(
            subcategory=sc, purchase_order__is_draft=False
        ).values("unit_price", "tot_amt", "tot_qty")
    )

    total_spent_amount = sum(float(i["tot_amt"]) for i in all_items)
    total_spent_quantity = sum(int(i["tot_qty"]) for i in all_items)

    total_ranges = len(price_ranges)

    # Build per-range accumulators in Python — O(items × ranges), negligible
    range_amt = {r.id: Decimal("0") for r in price_ranges}
    range_qty = {r.id: 0 for r in price_ranges}

    for item in all_items:
        unit_price = item["unit_price"]
        for r in price_ranges:
            buying_not_configured = (
                r.buying_from_range == Decimal("0") and r.buying_to_range == Decimal("0")
            )
            if buying_not_configured:
                continue
            if total_ranges == 1:
                range_amt[r.id] += item["tot_amt"]
                range_qty[r.id] += item["tot_qty"]
                break
            if r.buying_from_range <= unit_price <= r.buying_to_range:
                range_amt[r.id] += item["tot_amt"]
                range_qty[r.id] += item["tot_qty"]
                break

    ranges_list = [
        {
            "id": r.id,
            "sales_from": float(r.sales_from_range),
            "sales_to": float(r.sales_to_range),
            "buying_from": float(r.buying_from_range),
            "buying_to": float(r.buying_to_range),
            "approved_amount": float(r.approved_amount),
            "approved_quantity": r.approved_quantity,
            "spent_amount": float(range_amt[r.id]),
            "spent_quantity": int(range_qty[r.id]),
        }
        for r in price_ranges
    ]

    return JsonResponse({
        "success": True,
        "ranges": ranges_list,
        "total_spent_amount": total_spent_amount,
        "total_spent_quantity": total_spent_quantity,
    })


# ---------------------------------------------------------------------------
# Size manager
# ---------------------------------------------------------------------------

@login_required
def size_manager(request):
    q = (request.GET.get("q") or "").strip()
    subcategories_qs = SubCategory.objects.filter(sizes__isnull=False).distinct().prefetch_related("sizes").order_by("name")
    if q:
        subcategories_qs = subcategories_qs.filter(name__icontains=q) | subcategories_qs.filter(ch4_code__icontains=q)
    
    paginator = Paginator(subcategories_qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    
    return render(request, "po_sheet/size_manager.html", {
        "page_obj": page_obj,
        "subcategories": page_obj.object_list,
        "search_query": q,
    })


@login_required
def get_subcategory_sizes(request, subcategory_id):
    try:
        sc = SubCategory.objects.get(id=subcategory_id)
    except (SubCategory.DoesNotExist, ValueError):
        sc = SubCategory.objects.filter(name=subcategory_id).first()
    if not sc:
        return JsonResponse({"success": True, "sizes": []})
    sizes = list(SubCategorySize.objects.filter(subcategory=sc).order_by("id").values("id", "name"))
    return JsonResponse({"success": True, "sizes": sizes})


@login_required
@require_POST
def add_subcategory_sizes(request):
    try:
        data = json.loads(request.body)
        subcat_id = data.get("subcategory_id")
        size_names = data.get("sizes", [])

        try:
            sc = SubCategory.objects.get(id=subcat_id)
        except (SubCategory.DoesNotExist, ValueError, TypeError):
            sc, _ = SubCategory.objects.get_or_create(name=str(subcat_id))

        for name in size_names:
            name = name.strip()
            if name:
                SubCategorySize.objects.get_or_create(subcategory=sc, name=name)

        sizes = list(SubCategorySize.objects.filter(subcategory=sc).order_by("id").values("id", "name"))
        return JsonResponse({"success": True, "sizes": sizes})
    except Exception:
        logger.exception("add_subcategory_sizes error")
        return JsonResponse({"success": False, "error": "Server error"})


@login_required
@require_POST
def remove_subcategory_size(request):
    try:
        data = json.loads(request.body)
        subcat_id = data.get("subcategory_id")
        size_id = data.get("size_id")

        try:
            sc = SubCategory.objects.get(id=subcat_id)
        except (SubCategory.DoesNotExist, ValueError, TypeError):
            sc = SubCategory.objects.filter(name=str(subcat_id)).first()
        if not sc:
            return JsonResponse({"success": False, "error": "Subcategory not found"})

        get_object_or_404(SubCategorySize, id=size_id, subcategory=sc).delete()
        sizes = list(SubCategorySize.objects.filter(subcategory=sc).order_by("id").values("id", "name"))
        return JsonResponse({"success": True, "sizes": sizes})
    except Exception:
        logger.exception("remove_subcategory_size error")
        return JsonResponse({"success": False, "error": "Server error"})


# ---------------------------------------------------------------------------
# User management (staff only)
# ---------------------------------------------------------------------------

@login_required
def manage_users(request):
    if not request.user.is_staff:
        messages.error(request, "Only administrators can manage users")
        return redirect("po_sheet")

    if request.method == "POST":
        action = request.POST.get("action")
        username = request.POST.get("username", "").strip()
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")
        is_staff = request.POST.get("role") == "admin"
        is_active = "is_active" in request.POST

        if action == "create":
            if not username or not password:
                messages.error(request, "Username and Password are required")
            elif User.objects.filter(username=username).exists():
                messages.error(request, f"User '{username}' already exists")
            else:
                try:
                    User.objects.create_user(
                        username=username, password=password, email=email,
                        first_name=first_name, last_name=last_name,
                        is_staff=is_staff, is_active=is_active,
                    )
                    messages.success(request, f"User '{username}' created")
                except Exception as e:
                    messages.error(request, f"Error: {e}")
            return redirect("manage_users")

        if action == "edit":
            user_to_edit = get_object_or_404(User, id=request.POST.get("user_id"))
            if username and username != user_to_edit.username:
                if User.objects.filter(username=username).exclude(id=user_to_edit.id).exists():
                    messages.error(request, f"Username '{username}' already taken")
                    return redirect("manage_users")
                user_to_edit.username = username
            user_to_edit.first_name = first_name
            user_to_edit.last_name = last_name
            user_to_edit.email = email
            user_to_edit.is_staff = is_staff
            user_to_edit.is_active = is_active
            if password:
                user_to_edit.set_password(password)
            try:
                user_to_edit.save()
                messages.success(request, f"User '{user_to_edit.username}' updated")
            except Exception as e:
                messages.error(request, f"Error: {e}")
            return redirect("manage_users")

    users = User.objects.all().order_by("-date_joined")
    return render(request, "po_sheet/manage_users.html", {"users": users})


@login_required
def delete_user(request, user_id):
    if not request.user.is_staff:
        messages.error(request, "Only administrators can manage users")
        return redirect("po_sheet")
    user_to_delete = get_object_or_404(User, id=user_id)
    if user_to_delete == request.user:
        messages.error(request, "You cannot delete your own account")
    else:
        username = user_to_delete.username
        user_to_delete.delete()
        messages.success(request, f"User '{username}' deleted")
    return redirect("manage_users")


# ---------------------------------------------------------------------------
# Excel Upload Views
# ---------------------------------------------------------------------------

@login_required
def upload_excel(request):
    if not request.user.is_staff:
        messages.error(request, "Only admin can upload data")
        return redirect("po_sheet")

    context = {"results": None, "upload_type": None}

    if request.method == "POST":
        upload_type = request.POST.get("upload_type", "")
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "No file uploaded.")
            return render(request, "po_sheet/upload_excel.html", context)

        if not excel_file.name.endswith((".xlsx", ".xls")):
            messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
            return render(request, "po_sheet/upload_excel.html", context)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            if upload_type == "price_range":
                context["results"] = _process_price_range_upload(ws, request.user)
                context["upload_type"] = "price_range"
            elif upload_type == "buyer":
                resync = request.POST.get("resync") == "1"
                context["results"] = _process_buyer_upload(ws, resync=resync)
                context["upload_type"] = "buyer"
            else:
                messages.error(request, "Invalid upload type.")

        except Exception as e:
            logger.error(f"Excel upload error: {e}")
            messages.error(request, f"Error processing file: {e}")

    return render(request, "po_sheet/upload_excel.html", context)


def _process_price_range_upload(ws, user):
    """
    Expected columns (row 1 = header):
    SubCategory Name | SubCategory Code | Buyer | Sales From | Sales To | Buying From | Buying To | Approved Budget | Approved Qty
    """
    results = {"created": 0, "updated": 0, "skipped": 0, "errors": [], "rows": 0}

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, name_variants):
        for variant in name_variants:
            for i, h in enumerate(headers):
                if variant in h:
                    v = row[i].value
                    return v
        return None

    for row in ws.iter_rows(min_row=2, values_only=False):
        if all(c.value is None for c in row):
            continue
        results["rows"] += 1
        row_num = row[0].row
        try:
            subcat_name = str(col(row, ["subcategory name", "subcategory", "sub category", "name"])).strip() if col(row, ["subcategory name", "subcategory", "sub category", "name"]) else ""
            subcat_code_val = str(col(row, ["subcategory code", "subcat code", "code", "ch4"])).strip() if col(row, ["subcategory code", "subcat code", "code", "ch4"]) else ""
            buyer_name = str(col(row, ["buyer"])).strip() if col(row, ["buyer"]) else ""

            sales_from = col(row, ["sales from", "sale from"])
            sales_to   = col(row, ["sales to",   "sale to"])
            buying_from = col(row, ["buying from", "buy from"])
            buying_to   = col(row, ["buying to",   "buy to"])
            approved_budget = col(row, ["approved budget", "budget", "approved amount"])
            approved_qty    = col(row, ["approved qty", "quantity", "qty"])

            if not subcat_name or subcat_name.lower() == "none":
                results["errors"].append(f"Row {row_num}: Missing subcategory name")
                continue

            # Find or create subcategory
            sc = None
            if subcat_code_val and subcat_code_val.lower() != "none":
                sc = SubCategory.objects.filter(ch4_code=subcat_code_val).first()
            if not sc:
                sc, created = SubCategory.objects.get_or_create(
                    name=subcat_name,
                    defaults={"ch4_code": subcat_code_val or None}
                )
                if not created and subcat_code_val and subcat_code_val.lower() != "none" and not sc.ch4_code:
                    sc.ch4_code = subcat_code_val
                    sc.save(update_fields=["ch4_code"])

            # Link buyer if provided
            if buyer_name and buyer_name.lower() != "none":
                buyer_obj, _ = Buyer.objects.get_or_create(name=buyer_name)
                sc.buyers.add(buyer_obj)

            # Build field values
            sf = Decimal(str(sales_from or 0))
            st = Decimal(str(sales_to or 0))
            bf = Decimal(str(buying_from or 0))
            bt = Decimal(str(buying_to or 0))
            ab = Decimal(str(approved_budget or 0))
            aq = int(approved_qty or 0)

            # Upsert price range
            existing = SubCategoryPriceRange.objects.filter(
                subcategory=sc,
                sales_from_range=sf,
                sales_to_range=st,
            ).first()

            sc_code_for_range = subcat_code_val if subcat_code_val and subcat_code_val.lower() != "none" else None

            if existing:
                changed = (
                    existing.buying_from_range != bf or
                    existing.buying_to_range   != bt or
                    existing.approved_amount   != ab or
                    existing.approved_quantity != aq or
                    (sc_code_for_range and existing.subcat_code != sc_code_for_range)
                )
                if changed:
                    existing.buying_from_range = bf
                    existing.buying_to_range   = bt
                    existing.approved_amount   = ab
                    existing.approved_quantity = aq
                    if sc_code_for_range:
                        existing.subcat_code = sc_code_for_range
                    existing.save()
                    results["updated"] += 1
                else:
                    results["skipped"] += 1
            else:
                SubCategoryPriceRange.objects.create(
                    subcategory=sc,
                    sales_from_range=sf,
                    sales_to_range=st,
                    buying_from_range=bf,
                    buying_to_range=bt,
                    approved_amount=ab,
                    approved_quantity=aq,
                    subcat_code=sc_code_for_range,
                )
                results["created"] += 1

        except Exception as e:
            results["errors"].append(f"Row {row_num}: {e}")

    return results


def _process_buyer_upload(ws, resync=False):
    """
    Expected columns: SubCategory | SubCateCode | PGR | TVM Buyer
    Same format as buyer.xlsx.
    resync=True clears all existing buyer links before importing.
    """
    results = {"created": 0, "updated": 0, "skipped": 0, "no_buyer": 0, "errors": [], "rows": 0}

    if resync:
        SubCategory.buyers.through.objects.all().delete()

    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        results["rows"] += 1
        try:
            subcat_name = str(row[0]).strip() if row[0] else ""
            subcat_code = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            raw_buyer   = row[3] if len(row) > 3 else None
            buyer_name  = str(raw_buyer).strip() if raw_buyer and str(raw_buyer).strip() not in ("0", "") else ""

            if not subcat_name:
                results["errors"].append(f"Row {results['rows']+1}: Missing subcategory name")
                continue

            # Skip duplicate (subcat, buyer) pairs in same file
            key = (subcat_name, buyer_name)
            if key in seen:
                results["skipped"] += 1
                continue
            seen.add(key)

            # Always upsert the subcategory
            sc, sc_created = SubCategory.objects.get_or_create(
                name=subcat_name,
                defaults={"ch4_code": subcat_code or None}
            )
            if not sc_created and subcat_code and sc.ch4_code != subcat_code:
                sc.ch4_code = subcat_code
                sc.save(update_fields=["ch4_code"])
                results["updated"] += 1

            # Link buyer only when a real name exists
            if buyer_name:
                buyer_obj, _ = Buyer.objects.get_or_create(name=buyer_name)
                if not sc.buyers.filter(id=buyer_obj.id).exists():
                    sc.buyers.add(buyer_obj)
                    results["created"] += 1
                else:
                    results["skipped"] += 1
            else:
                results["no_buyer"] += 1

        except Exception as e:
            results["errors"].append(f"Row {results['rows']+1}: {e}")

    return results
